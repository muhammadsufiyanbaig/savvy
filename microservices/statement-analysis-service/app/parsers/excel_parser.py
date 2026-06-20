import io
import logging
from typing import Dict, List, Optional

from app.parsers.base_parser import BaseParser
from app.utils.exceptions import ParsingError

logger = logging.getLogger(__name__)

_DATE_KEYS = {"date", "transaction date", "posted date"}
_DESC_KEYS = {"description", "merchant", "details", "memo", "payee"}
_AMOUNT_KEYS = {"amount", "debit", "credit", "transaction amount"}


class ExcelParser(BaseParser):
    """Parse bank statements from Excel (.xlsx / .xls) files."""

    def parse(self, file_content: bytes) -> Dict:
        try:
            import openpyxl  # lazy import
        except ImportError as exc:
            raise ParsingError("openpyxl not installed. Run: pip install openpyxl") from exc

        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            sheet = self._find_transaction_sheet(wb)
            transactions = self._extract_rows(sheet)

            return {
                "text": "",
                "transactions": transactions,
                "row_count": len(transactions),
                "metadata": {
                    "sheet_name": sheet.title,
                    "sheet_count": len(wb.worksheets),
                },
            }
        except ParsingError:
            raise
        except Exception as exc:
            logger.error("Excel parse error: %s", exc)
            raise ParsingError(f"Excel parsing failed: {exc}") from exc

    # ── private ───────────────────────────────────────────────────────────────

    @staticmethod
    def _find_transaction_sheet(wb):
        """Return first sheet whose header row contains date/description/amount keywords."""
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(min_row=1, max_row=3, values_only=True):
                if row and any(
                    str(cell).lower().strip() in _DATE_KEYS | _DESC_KEYS | _AMOUNT_KEYS
                    for cell in row
                    if cell is not None
                ):
                    return sheet
        return wb.worksheets[0]

    def _extract_rows(self, sheet) -> List[Dict]:
        rows_iter = sheet.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if header_row is None:
            return []

        headers = [str(h).lower().strip() if h else "" for h in header_row]

        def _col(keys):
            for k in keys:
                if k in headers:
                    return headers.index(k)
            return None

        date_idx = _col(_DATE_KEYS)
        desc_idx = _col(_DESC_KEYS)
        amt_idx = _col(_AMOUNT_KEYS)

        transactions: List[Dict] = []
        for row in rows_iter:
            if all(cell is None for cell in row):
                continue
            txn = {
                "date": str(row[date_idx]).strip() if date_idx is not None and row[date_idx] else "",
                "description": str(row[desc_idx]).strip() if desc_idx is not None and row[desc_idx] else "",
                "amount": self._clean_amount(row[amt_idx]) if amt_idx is not None else 0.0,
                "transaction_type": "debit",
            }
            if txn["description"] and txn["amount"] > 0:
                transactions.append(txn)

        return transactions
